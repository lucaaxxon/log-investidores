"""
Sincroniza entradas novas da nuvem com a planilha local.
Roda ao dar duplo clique em sincronizar.bat
"""
import requests
import openpyxl
import json
import os
from datetime import datetime

# ── Configuração ────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "sync_config.json")
STATE_FILE  = os.path.join(BASE_DIR, "sync_state.json")
EXCEL_FILE  = os.path.join(BASE_DIR, "2026 06 23 LPs - Contato_v12.xlsx")
# ────────────────────────────────────────────────────────────────────────────


def load_config():
    if not os.path.exists(CONFIG_FILE):
        print("\nArquivo sync_config.json não encontrado.")
        print("Crie o arquivo com o conteúdo:")
        print('  {"url": "https://seu-app.railway.app"}')
        input("\nPressione Enter para sair.")
        raise SystemExit(1)
    with open(CONFIG_FILE, encoding="utf-8-sig") as f:
        cfg = json.load(f)
    if not cfg.get("url"):
        print("URL não configurada em sync_config.json")
        input("Pressione Enter para sair.")
        raise SystemExit(1)
    return cfg


def load_state():
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"last_id": 0}


def save_state(state):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f)


def get_existing_investors(ws_tabela):
    existing = set()
    for row in ws_tabela.iter_rows(min_row=28, values_only=True):
        val = row[1]
        if val and isinstance(val, str) and val.strip():
            existing.add(val.strip().lower())
    return existing


def get_last_tabela_row(ws_tabela):
    last = 27
    for row in ws_tabela.iter_rows(min_row=28):
        if row[1].value:
            last = row[0].row
    return last


def parse_date(s):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return s  # fallback: store as string


def main():
    print("=" * 50)
    print("  Sincronização com a nuvem")
    print("=" * 50)

    cfg   = load_config()
    state = load_state()
    url   = cfg["url"].rstrip("/")

    print(f"\nBuscando entradas novas (desde id={state['last_id']})...")

    try:
        resp = requests.get(f"{url}/export?since_id={state['last_id']}", timeout=15)
        resp.raise_for_status()
        data = resp.json()
    except Exception as e:
        print(f"\nErro ao conectar com a nuvem: {e}")
        input("\nPressione Enter para sair.")
        return

    rows      = data.get("rows", [])
    investors = data.get("investors", [])
    last_id   = data.get("last_id", state["last_id"])

    if not rows:
        print("\nNenhuma entrada nova para sincronizar.")
        input("\nPressione Enter para sair.")
        return

    print(f"  {len(rows)} entrada(s) nova(s) encontrada(s).")

    # Abre a planilha local
    if not os.path.exists(EXCEL_FILE):
        print(f"\nArquivo Excel não encontrado:\n  {EXCEL_FILE}")
        input("\nPressione Enter para sair.")
        return

    wb         = openpyxl.load_workbook(EXCEL_FILE)
    ws_log     = wb["Log"]
    ws_tabela  = wb["Tabela"]

    # Adiciona investidores novos na aba Tabela
    existing_investors = get_existing_investors(ws_tabela)
    added_investors    = []
    for nome in investors:
        if nome.strip().lower() not in existing_investors:
            last_row = get_last_tabela_row(ws_tabela)
            ws_tabela.cell(row=last_row + 1, column=2, value=nome.strip())
            existing_investors.add(nome.strip().lower())
            added_investors.append(nome)

    # Adiciona linhas de log
    for row in rows:
        ws_log.append([
            None,
            row.get("investidor", ""),
            parse_date(row.get("data", "")),
            row.get("tipo_contato", ""),
            row.get("apresentou_arquivo", ""),
            parse_date(row.get("follow_up", "")) if row.get("follow_up") else None,
            row.get("comentario", ""),
            row.get("fase_pipeline", ""),
            row.get("motivo_descarte", ""),
        ])

    wb.save(EXCEL_FILE)
    state["last_id"] = last_id
    save_state(state)

    print(f"\n  Adicionadas {len(rows)} linha(s) na aba Log.")
    if added_investors:
        print(f"  Adicionados {len(added_investors)} investidor(es) novo(s) na aba Tabela:")
        for nome in added_investors:
            print(f"    - {nome}")
    print("\nSincronização concluída com sucesso!")
    input("\nPressione Enter para sair.")


if __name__ == "__main__":
    main()
